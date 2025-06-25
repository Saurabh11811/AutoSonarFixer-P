#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import logging
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook

# Import config manager
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(project_root, "utils"))
from config_manager import ConfigManager

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

global_stats = {}

def process_full_snapshot_files(config):
    sonar_config = config['sonarqube']
    results_path = sonar_config['results_path']
    repos = config['github']['repos']

    for repo in repos:
        if not repo.get('enabled', False): continue

        repo_name = repo['repo_url'].rstrip('/').split('/')[-1].replace('.git', '')
        repo_results_dir = os.path.join(results_path, repo_name)

        latest_file = None
        files = [f for f in os.listdir(repo_results_dir) if f.endswith("_full_snapshot.json")]
        files.sort(reverse=True)
        if files: latest_file = files[0]

        if not latest_file:
            logging.warning(f"No full_snapshot found for repo {repo_name}")
            continue

        file_path = os.path.join(repo_results_dir, latest_file)
        with open(file_path, 'r') as f:
            raw_data = json.load(f)

        issues = raw_data.get("issues", [])
        hotspots = raw_data.get("hotspots", [])

        # For normalization output
        normalized_issues = []
        file_issue_map = defaultdict(lambda: defaultdict(int))
        maintainability = reliability = security = other = 0
        issues_count = 0

        # Normalize Issues
        for issue in issues:
            component = issue.get("component", "")
            file_path_str = component.split(":", 1)[-1]
            issue_type = issue.get('type')

            norm_issue = {
                "file": file_path_str,
                "line": issue.get("line"),
                "rule": issue.get("rule"),
                "severity": issue.get("severity"),
                "message": issue.get("message"),
                "type": issue_type,
                "source": "issues",
                "impacts": issue.get("impacts", [])
            }
            normalized_issues.append(norm_issue)

            file_issue_map[file_path_str][issue_type] += 1
            issues_count += 1

            for impact in issue.get("impacts", []):
                quality = impact.get("softwareQuality", "")
                if quality == "MAINTAINABILITY": maintainability += 1
                elif quality == "RELIABILITY": reliability += 1
                elif quality == "SECURITY": security += 1
                else: other += 1

        # Normalize Hotspots
        for hotspot in hotspots:
            component = hotspot.get("component", "")
            file_path_str = component.split(":", 1)[-1]

            norm_hotspot = {
                "file": file_path_str,
                "line": hotspot.get("line"),
                "rule": hotspot.get("ruleKey"),
                "severity": hotspot.get("vulnerabilityProbability"),
                "message": hotspot.get("message"),
                "type": "SECURITY_HOTSPOT",
                "source": "hotspots"
            }
            normalized_issues.append(norm_hotspot)
            file_issue_map[file_path_str]["SECURITY_HOTSPOT"] += 1

        # Store normalized issues file
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        normalized_file_path = os.path.join(repo_results_dir, f"{timestamp}_normalized_issues.json")
        with open(normalized_file_path, 'w') as f_out:
            json.dump(normalized_issues, f_out, indent=2)
        logging.info(f"✅ Normalized issues stored at: {normalized_file_path}")

        # Store aggregated stats for console & excel
        global_stats[repo_name] = {
            "total_issues": issues_count,
            "maintainability": maintainability,
            "reliability": reliability,
            "security": security,
            "other": other,
            "hotspots": len(hotspots),
            "files": file_issue_map
        }

def print_console_summary():
    logging.info("\n=========== SONAR SUMMARY ===========")
    for repo, stats in global_stats.items():
        logging.info(f"\nRepo: {repo}")
        logging.info(f"  Total Issues: {stats['total_issues']}")
        logging.info(f"  Maintainability: {stats['maintainability']}")
        logging.info(f"  Reliability: {stats['reliability']}")
        logging.info(f"  Security (issues): {stats['security']}")
        logging.info(f"  Other impacts: {stats['other']}")
        logging.info(f"  Security Hotspots: {stats['hotspots']}")
        logging.info(f"  Unique Files with Issues: {len(stats['files'])}")
    logging.info("\n=====================================")

def write_excel_report(config):
    sonar_config = config['sonarqube']
    results_path = sonar_config['results_path']
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_path = os.path.join(results_path, f"sonar_file_summary_{timestamp}.xlsx")

    wb = Workbook()
    wb.active.title = "Temporary"

    for repo, stats in global_stats.items():
        ws_overview = wb.create_sheet(title=f"{repo}-O")
        ws_overview.append(["Metric", "Value"])
        ws_overview.append(["Total Issues", stats['total_issues']])
        ws_overview.append(["Maintainability", stats['maintainability']])
        ws_overview.append(["Reliability", stats['reliability']])
        ws_overview.append(["Security", stats['security']])
        ws_overview.append(["Other impacts", stats['other']])
        ws_overview.append(["Security Hotspots", stats['hotspots']])
        ws_overview.append(["Unique Files with Issues", len(stats['files'])])

        ws_details = wb.create_sheet(title=f"{repo}-D")
        headers = ["File Path", "File Name", "CODE_SMELL", "BUG", "VULNERABILITY", "SECURITY_HOTSPOT", "TOTAL"]
        ws_details.append(headers)

        sorted_files = sorted(stats['files'].items(), key=lambda x: (x[0], os.path.basename(x[0])))
        totals = {"CODE_SMELL":0, "BUG":0, "VULNERABILITY":0, "SECURITY_HOTSPOT":0, "TOTAL":0}

        for file_path, issue_types in sorted_files:
            file_name = os.path.basename(file_path)
            cs = issue_types.get("CODE_SMELL",0)
            bug = issue_types.get("BUG",0)
            vuln = issue_types.get("VULNERABILITY",0)
            hs = issue_types.get("SECURITY_HOTSPOT",0)
            total = cs + bug + vuln + hs
            ws_details.append([file_path, file_name, cs, bug, vuln, hs, total])

            totals["CODE_SMELL"] += cs
            totals["BUG"] += bug
            totals["VULNERABILITY"] += vuln
            totals["SECURITY_HOTSPOT"] += hs
            totals["TOTAL"] += total

        ws_details.append(["TOTAL", "", totals["CODE_SMELL"], totals["BUG"], totals["VULNERABILITY"], totals["SECURITY_HOTSPOT"], totals["TOTAL"]])

    if "Temporary" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Temporary"])

    wb.save(excel_path)
    logging.info(f"\n✅ Excel report generated: {excel_path}")

def run_summary():
    config_mgr = ConfigManager()
    config = config_mgr.config

    process_full_snapshot_files(config)
    print_console_summary()
    write_excel_report(config)

if __name__ == "__main__":
    run_summary()
